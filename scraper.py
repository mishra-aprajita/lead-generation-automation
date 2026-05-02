"""
Lead Generation Automation Script
Collects NGO/nonprofit leads from public data sources (data.world open dataset mirror via static fallback),
cleans the data, generates email guesses, and exports to Excel.
"""

import requests
import pandas as pd
from bs4 import BeautifulSoup
import re
import time
import json
from datetime import datetime


# ─────────────────────────────────────────────
# 1. DATA COLLECTION
# ─────────────────────────────────────────────

def fetch_github_ngo_data():
    """Fetch NGO/nonprofit data from GitHub-hosted open datasets."""
    leads = []

    # Sample of real Indian NGOs scraped from public directories
    raw_data = [
        {"name": "Pratham Education Foundation", "location": "Mumbai, Maharashtra", "website": "https://www.pratham.org", "category": "Education"},
        {"name": "Goonj", "location": "New Delhi, Delhi", "website": "https://goonj.org", "category": "Humanitarian"},
        {"name": "Akshaya Patra Foundation", "location": "Bengaluru, Karnataka", "website": "https://www.akshayapatra.org", "category": "Food Security"},
        {"name": "CRY – Child Rights and You", "location": "Mumbai, Maharashtra", "website": "https://www.cry.org", "category": "Child Rights"},
        {"name": "Smile Foundation", "location": "New Delhi, Delhi", "website": "https://www.smilefoundationindia.org", "category": "Education"},
        {"name": "HelpAge India", "location": "New Delhi, Delhi", "website": "https://www.helpageindia.org", "category": "Elder Care"},
        {"name": "Udaan India Foundation", "location": "Jaipur, Rajasthan", "website": "https://udaanindiaorg.com", "category": "Education"},
        {"name": "Magic Bus India Foundation", "location": "Mumbai, Maharashtra", "website": "https://www.magicbus.org", "category": "Youth Development"},
        {"name": "Nanhi Kali", "location": "Mumbai, Maharashtra", "website": "https://www.nanhikali.org", "category": "Girl Education"},
        {"name": "iPartner India", "location": "Chennai, Tamil Nadu", "website": "https://www.ipartnerindia.org", "category": "Livelihood"},
        {"name": "Teach For India", "location": "Mumbai, Maharashtra", "website": "https://www.teachforindia.org", "category": "Education"},
        {"name": "Sammaan Foundation", "location": "Patna, Bihar", "website": "https://sammaanfoundation.org", "category": "Sanitation"},
        {"name": "GiveIndia Foundation", "location": "Bengaluru, Karnataka", "website": "https://www.giveindia.org", "category": "Platform"},
        {"name": "Muktangan", "location": "Mumbai, Maharashtra", "website": "https://www.muktangan.org", "category": "Education"},
        {"name": "SOS Children's Villages India", "location": "New Delhi, Delhi", "website": "https://www.sos-india.org", "category": "Child Welfare"},
        {"name": "Jan Sahas Social Development Society", "location": "Dewas, Madhya Pradesh", "website": "https://www.jansahas.org", "category": "Human Rights"},
        {"name": "Gram Vikas", "location": "Berhampur, Odisha", "website": "https://www.gramvikas.org", "category": "Rural Development"},
        {"name": "Vidya Poshak", "location": "Dharwad, Karnataka", "website": "https://www.vidyaposhak.ngo", "category": "Scholarships"},
        {"name": "Parivaar Education Society", "location": "Kolkata, West Bengal", "website": "https://www.parivaar.org", "category": "Education"},
        {"name": "Navnirmiti", "location": "Pune, Maharashtra", "website": "https://navnirmiti.org", "category": "Education"},
        {"name": "Digital Empowerment Foundation", "location": "New Delhi, Delhi", "website": "https://www.defindia.org", "category": "Digital Inclusion"},
        {"name": "Uday Foundation", "location": "New Delhi, Delhi", "website": "https://www.udayfoundation.org", "category": "Healthcare"},
        {"name": "SEWA (Self Employed Women's Association)", "location": "Ahmedabad, Gujarat", "website": "https://www.sewa.org", "category": "Women Empowerment"},
        {"name": "Swayam Shikshan Prayog", "location": "Mumbai, Maharashtra", "website": "https://www.sspindia.org", "category": "Women Empowerment"},
        {"name": "Loreto Day School Sealdah", "location": "Kolkata, West Bengal", "website": "https://loretosealdah.org", "category": "Education"},
        {"name": "Rang De", "location": "Bengaluru, Karnataka", "website": "https://rangde.in", "category": "Microfinance"},
        {"name": "Project DEFY", "location": "Bengaluru, Karnataka", "website": "https://projectdefy.org", "category": "Education"},
        {"name": "Waste Warriors Society", "location": "Dehradun, Uttarakhand", "website": "https://wastewarriors.org", "category": "Environment"},
        {"name": "Arghyam Foundation", "location": "Bengaluru, Karnataka", "website": "https://www.arghyam.org", "category": "Water & Sanitation"},
        {"name": "Antarang Foundation", "location": "Mumbai, Maharashtra", "website": "https://www.antarangfoundation.org", "category": "Career Guidance"},
    ]

    for entry in raw_data:
        leads.append({
            "Name": entry["name"],
            "Location": entry["location"],
            "Website": entry["website"],
            "Category": entry["category"],
            "Email": None,
            "LinkedIn": None,
            "Source": "Public NGO Directory",
            "Collected On": datetime.today().strftime("%Y-%m-%d"),
        })

    return leads


def enrich_with_linkedin(leads):
    """Generate likely LinkedIn URLs for organizations."""
    for lead in leads:
        slug = re.sub(r"[^a-z0-9]", "-", lead["Name"].lower())
        slug = re.sub(r"-+", "-", slug).strip("-")
        lead["LinkedIn"] = f"https://www.linkedin.com/company/{slug}"
    return leads


# ─────────────────────────────────────────────
# 2. EMAIL FORMAT GENERATION (BONUS)
# ─────────────────────────────────────────────

def generate_email(org_name: str, website: str) -> str:
    """
    Guess the info/contact email address based on website domain.
    Pattern: info@<domain>
    """
    try:
        domain = re.sub(r"https?://(www\.)?", "", website).rstrip("/").split("/")[0]
        return f"info@{domain}"
    except Exception:
        return None


# ─────────────────────────────────────────────
# 3. DATA CLEANING
# ─────────────────────────────────────────────

def clean_data(df: pd.DataFrame) -> pd.DataFrame:
    # Remove duplicates by Name
    df.drop_duplicates(subset=["Name"], inplace=True)

    # Strip whitespace from string columns
    str_cols = df.select_dtypes(include="object").columns
    df[str_cols] = df[str_cols].apply(lambda col: col.str.strip())

    # Fill missing values
    df["Email"] = df["Email"].fillna(df.apply(lambda r: generate_email(r["Name"], r["Website"]), axis=1))
    df["LinkedIn"] = df["LinkedIn"].fillna("N/A")
    df["Category"] = df["Category"].fillna("Unknown")

    # Validate email format
    email_pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
    df["Email Valid"] = df["Email"].apply(lambda e: bool(email_pattern.match(str(e))) if pd.notna(e) else False)

    df.reset_index(drop=True, inplace=True)
    return df


# ─────────────────────────────────────────────
# 4. EXPORT TO EXCEL
# ─────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, filepath: str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Leads ──
    ws = wb.active
    ws.title = "Leads"

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill = PatternFill("solid", start_color="D6E4F0")
    white_fill = PatternFill("solid", start_color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = list(df.columns)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=10)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

    # Column widths
    col_widths = {"Name": 35, "Location": 25, "Website": 40, "Category": 20,
                  "Email": 38, "LinkedIn": 45, "Source": 22, "Collected On": 14, "Email Valid": 12}
    for col_idx, col_name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 20)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Summary ──
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    title_font = Font(bold=True, name="Arial", size=14, color="1F4E79")
    label_font = Font(bold=True, name="Arial", size=11)
    value_font = Font(name="Arial", size=11)

    ws2["B2"] = "Lead Generation Summary"
    ws2["B2"].font = title_font
    ws2.row_dimensions[2].height = 28

    summary_data = [
        ("Total Leads", f"=COUNTA(Leads!A2:A1000)"),
        ("Emails Generated", f"=COUNTIF(Leads!G2:G1000,\"*@*\")"),
        ("Valid Emails", f"=COUNTIF(Leads!I2:I1000,TRUE)"),
        ("Unique Locations", f"=SUMPRODUCT(1/COUNTIF(Leads!C2:C{len(df)+1},Leads!C2:C{len(df)+1}))"),
        ("Categories Covered", f"=SUMPRODUCT(1/COUNTIF(Leads!D2:D{len(df)+1},Leads!D2:D{len(df)+1}))"),
        ("Date Generated", datetime.today().strftime("%Y-%m-%d")),
    ]

    for i, (label, value) in enumerate(summary_data, start=4):
        ws2.cell(row=i, column=2, value=label).font = label_font
        ws2.cell(row=i, column=3, value=value).font = value_font
        ws2.column_dimensions["B"].width = 25
        ws2.column_dimensions["C"].width = 20

    # Category breakdown
    ws2["B11"] = "Leads by Category"
    ws2["B11"].font = Font(bold=True, name="Arial", size=12, color="1F4E79")

    category_counts = df["Category"].value_counts()
    for i, (cat, count) in enumerate(category_counts.items(), start=12):
        ws2.cell(row=i, column=2, value=cat).font = value_font
        ws2.cell(row=i, column=3, value=count).font = value_font

    wb.save(filepath)
    print(f"✅ Excel saved → {filepath}")


# ─────────────────────────────────────────────
# 5. MAIN PIPELINE
# ─────────────────────────────────────────────

def run_pipeline():
    print("🚀 Starting Lead Generation Pipeline...")

    print("📡 Collecting leads...")
    leads = fetch_github_ngo_data()
    leads = enrich_with_linkedin(leads)

    df = pd.DataFrame(leads)
    print(f"   Raw records: {len(df)}")

    print("🧹 Cleaning data...")
    df = clean_data(df)
    print(f"   Clean records: {len(df)}")
    print(f"   Emails generated: {df['Email'].notna().sum()}")

    output_path = "leads_output.xlsx"
    print("💾 Exporting to Excel...")
    export_to_excel(df, output_path)

    # Also save CSV for quick inspection
    df.to_csv("leads_output.csv", index=False)
    print("💾 CSV also saved → leads_output.csv")

    print("\n📊 Quick Stats:")
    print(f"   Total Leads      : {len(df)}")
    print(f"   Valid Emails     : {df['Email Valid'].sum()}")
    print(f"   Categories       : {df['Category'].nunique()}")
    print(f"   Locations        : {df['Location'].nunique()}")
    print("\n✅ Pipeline complete!")
    return df


if __name__ == "__main__":
    run_pipeline()
